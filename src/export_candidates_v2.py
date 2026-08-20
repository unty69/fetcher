"""
export_candidates_v2.py — Step 4 production candidate export.

Scope: the FULL is_alt_cluster population (2+ phone OR 2+ username, excluding
the 4 suspected_non_discovery mega-clusters) -- NOT alt_cluster_review.py's
select_candidates() (username_count>=2 only), which validate_v2.py showed
returns exactly the already-labeled 135 -- a dead end for export. This script
deliberately uses the broader definition.

Excludes all canonical_id already present in the root ground-truth workbook
(any вердикт, including the blank row -- 189 as of the last merge: 135
original + 54 newly labeled from this pipeline's own prior export), then
applies:

  HARD excludes (re-validated on the full 189-identity ground truth after
  each merge: 0 false positives on any known web identity) --
  has_military_flag, has_agency_self_id, NOT has_verified_relevant_message
  (category AND a CPA-verified employer brand matched on the SAME message
  text -- fixes an earlier identity-level union bug where categories and
  brands were each accumulated independently across ALL of an identity's
  messages), has_pvz_mention (NEW -- Арсений: "пункт выдачи"/ПВЗ roles are
  single/few-person points, not mass-hire targets; confirmed on real message
  text that PVZ postings get swept into "Курьеры и доставка" because the job
  description mentions receiving deliveries FROM a courier, not because the
  role is one -- full hard-exclude, not a soft signal, per instruction).

  SOFT signals only, used for ranking, never for exclusion -- employer_count
  (bucketed per the CLAUDE.md hypothesis: 2-3 up, 0 neutral, 1 or 4+ down;
  still unconfirmed at scale), content_volume_flag, has_ref_link,
  category_count (penalizes category dispersion above the survivor
  population's empirical top-tertile boundary), address_location_count
  (distinct street/metro/mall/bare-address matches across an identity's
  messages, deduped so the same address repeated across near-duplicate
  reposts doesn't inflate the count; validated against real message text --
  see validate_v2.py's ADDRESS_PATTERNS), off_target_fraction (NEW -- share
  of an identity's messages matching a curated off-target white-collar
  keyword list -- инженер/сметчик/тендер/аналитик/юрист/бухгалтер/
  маркетолог/ревизор/офис-менеджер/менеджер по продажам/директор -- AND
  matching no target category AND not already military/agency/PVZ; built
  and validated against real uncategorized message titles, deliberately
  narrow so plausibly-legitimate-but-uncovered content like "Кондитер"
  isn't swept in). All four penalty thresholds are the survivor
  population's own top-tertile boundary, not hardcoded guesses.

  channel_count/message_count are REMOVED from scoring (kept as reported
  columns only): diagnosed across the full 189-identity ground truth after
  the second merge and found NOT to cleanly separate web from non-web --
  heavy overlap across the whole range, with outliers in both classes (one
  known web identity has message_count=30). No signal was invented to
  replace them, per instruction.

  corp_domain is NOT part of the additive score -- it is a direct
  main-eligibility GATE: corp_domain=True identities are excluded from
  candidates_v2_main.xlsx regardless of score, but still land in
  candidates_v2_borderline.xlsx if they clear the export threshold (not
  hard-excluded from the export entirely).

Threshold: score > 0 (net positive evidence over negative) -- the natural
zero-crossing of a signed additive score, not a percentile picked to hit a
row-count target. Main/borderline split is by score-threshold PROXIMITY
(median split of the corp_domain=False exported population), not by
content_volume_flag -- that flag is boolean (see report) and CLAUDE.md
instructs falling back to score-proximity for the split when it's boolean
rather than tri-state. content_volume_flag still feeds the score itself as
one soft input.

Writes output/candidates_v2_main.xlsx and output/candidates_v2_borderline.xlsx.
Neither is committed (see .gitignore -- output/ is already ignored).
Never writes to cache/, never touches the ground-truth workbook (read_only).
"""
import statistics
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import alt_cluster_review as acr  # noqa: E402
import validate_v2 as v2  # noqa: E402 -- reuse collect_debug_evidence() for the
                           # has_verified_relevant_message co-occurrence examples
                           # needed in the sample-rows report below (single source
                           # of truth for the debug-snippet logic, not duplicated)

GROUND_TRUTH_PATH = BASE_DIR / "alt_clusters_review.xlsx"
OUTPUT_MAIN = BASE_DIR / "output" / "candidates_v2_main.xlsx"
OUTPUT_BORDERLINE = BASE_DIR / "output" / "candidates_v2_borderline.xlsx"


def load_ground_truth_cids():
    """All canonical_id present in the root workbook, regardless of вердикт
    or blank status -- read_only, never modifies the file."""
    wb = openpyxl.load_workbook(GROUND_TRUTH_PATH, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter)
        cid_col = {h: i for i, h in enumerate(header)}["canonical_id"]
        return {row[cid_col] for row in rows_iter
                if row is not None and not all(v is None for v in row) and row[cid_col]}
    finally:
        wb.close()


def employer_bucket_score(employer_count):
    """CLAUDE.md hypothesis (still unconfirmed at scale -- soft rank only):
    0=weak/unknown (neutral), 1=possible negative, 2-3=possible positive,
    4+=possible aggregator/noise."""
    if employer_count in (2, 3):
        return 2
    if employer_count == 0:
        return 0
    return -1  # 1 or 4+


def tertile_boundaries(values):
    """Returns (q1, q2) tertile cut points of the ACTUAL population -- not a
    guessed cut point. Every current caller uses these only for a one-sided
    "> q2 is high, penalize" check (category_count, address_location_count,
    off_target_fraction); none currently need a 3-way bucket assignment.
    Degenerate case (population too small/uniform for statistics.quantiles)
    returns q1=q2=the single value."""
    clean = sorted(v for v in values if v is not None)
    if len(clean) < 3 or clean[0] == clean[-1]:
        only = clean[0] if clean else 0
        return only, only
    return statistics.quantiles(clean, n=3)


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    graph = acr.load_identity_graph()
    identities = graph["identities"]
    contact_to_canonical = graph["contact_to_canonical"]

    alt_cluster_pool = {
        cid for cid, ident in identities.items()
        if ident["is_alt_cluster"] and not ident["suspected_non_discovery"]
    }
    print(f"[1] Full is_alt_cluster non-suspND population: {len(alt_cluster_pool)}")

    gt_cids = load_ground_truth_cids()
    print(f"[2] Ground-truth canonical_ids to exclude (any вердикт incl. blank): {len(gt_cids)}")
    gt_in_pool = gt_cids & alt_cluster_pool
    print(f"    of which present in the alt-cluster pool: {len(gt_in_pool)}")

    export_pool = alt_cluster_pool - gt_cids
    print(f"[3] Export pool (alt-cluster minus already-reviewed): {len(export_pool)}")

    print("\nComputing evidence (full message text) for the export pool -- one pass over all dumps...")
    # v2.collect_debug_evidence() mirrors acr.collect_evidence() exactly (same
    # regex objects/functions) but also captures verified_relevant_examples --
    # the actual co-occurring (category, verified_brand, title) per identity,
    # needed for the sample-rows report below.
    evidence = v2.collect_debug_evidence(export_pool, contact_to_canonical)

    # --- report content_volume_flag / soft-signal distributions on the RAW
    # ~7047 population, before any hard exclusion (what was asked for) ---
    print("\n" + "=" * 78)
    print(f"Soft-signal distributions on the raw export pool (n={len(export_pool)}, "
          f"BEFORE hard excludes)")
    print("=" * 78)
    cv_true = sum(1 for cid in export_pool if evidence[cid]["max_text_len"] >= acr.CONTENT_VOLUME_MIN_CHARS)
    print(f"content_volume_flag (max_text_len >= {acr.CONTENT_VOLUME_MIN_CHARS}, boolean, "
          f"NOT tri-state -- see collect_evidence()/build_rows() in alt_cluster_review.py):")
    print(f"  True (sufficient):  {cv_true}/{len(export_pool)} ({100*cv_true/len(export_pool):.1f}%)")
    print(f"  False (low-content): {len(export_pool)-cv_true}/{len(export_pool)} "
          f"({100*(len(export_pool)-cv_true)/len(export_pool):.1f}%)")
    print("  This is meaningfully different from the 100% seen on the hand-picked 135 in Step 3 --")
    print("  that population was pre-selected by earlier review, this one is not.")

    max_len_values = sorted(evidence[cid]["max_text_len"] for cid in export_pool)
    if max_len_values:
        pcts = statistics.quantiles(max_len_values, n=10) if len(set(max_len_values)) > 1 else [max_len_values[0]] * 9
        print(f"\n  max_text_len percentiles (chars): p10={pcts[0]:.0f} p25={pcts[1]:.0f} "
              f"p50(median)={pcts[4]:.0f} p75={pcts[6]:.0f} p90={pcts[8]:.0f} max={max_len_values[-1]}")

    emp_dist = {}
    for cid in export_pool:
        n = len(evidence[cid]["brands"])
        b = str(n) if n < 4 else "4+"
        emp_dist[b] = emp_dist.get(b, 0) + 1
    print(f"\nemployer_count distribution: " +
          ", ".join(f"{b}={emp_dist.get(b, 0)}" for b in ["0", "1", "2", "3", "4+"]))

    # --- hard excludes ---
    print("\n" + "=" * 78)
    print("Hard excludes (co-occurrence-fixed gate; re-validated: 0 false positives on known web)")
    print("=" * 78)
    mil = {cid for cid in export_pool if evidence[cid]["has_military_flag"]}
    ag = {cid for cid in export_pool if evidence[cid]["has_agency_self_id"]}
    not_verified = {cid for cid in export_pool if not evidence[cid]["has_verified_relevant_message"]}
    pvz = {cid for cid in export_pool if evidence[cid]["has_pvz_mention"]}
    excluded = mil | ag | not_verified | pvz
    print(f"has_military_flag=True:                 {len(mil)}")
    print(f"has_agency_self_id=True:                {len(ag)}")
    print(f"NOT has_verified_relevant_message:      {len(not_verified)}")
    print(f"has_pvz_mention=True (NEW -- Арсений: ПВЗ is not a mass-hire target, full hard-exclude): {len(pvz)}")
    print(f"  overlap military & agency:                {len(mil & ag)}")
    print(f"  overlap military & not_verified:          {len(mil & not_verified)}")
    print(f"  overlap agency & not_verified:             {len(ag & not_verified)}")
    print(f"  overlap pvz & not_verified:                 {len(pvz & not_verified)} "
          f"(PVZ mentions usually already fail has_verified_relevant_message too, but not always)")
    print(f"  overlap all four:                          {len(mil & ag & not_verified & pvz)}")
    print(f"Union excluded (any of the four):           {len(excluded)}")

    survivors = export_pool - excluded
    print(f"\nSurviving after hard excludes: {len(survivors)}")
    print("NOTE: employer_count (raw brand-match count) is explicitly NOT required > 0 on its own, "
          "and does NOT independently rescue an identity that fails the gate -- the gate itself "
          "already requires a CPA-verified employer brand to co-occur with a category on the SAME "
          "message (the 'Яндекс ищет менеджера' trap: a bare category or bare employer match, or "
          "both from different messages, is not enough).")

    # --- soft scoring on survivors only ---
    print("\n" + "=" * 78)
    print(f"Soft-signal distributions on SURVIVORS (n={len(survivors)}, post hard-exclude, "
          f"the actual scoring population)")
    print("=" * 78)
    cv_true_surv = sum(1 for cid in survivors if evidence[cid]["max_text_len"] >= acr.CONTENT_VOLUME_MIN_CHARS)
    print(f"content_volume_flag True: {cv_true_surv}/{len(survivors)} "
          f"({100*cv_true_surv/len(survivors):.1f}%)" if survivors else "n/a")

    # channel_count/message_count: REMOVED from scoring this round. Diagnosed
    # across the full 189-identity ground truth (see conversation record):
    # web message_count = [1,2,3,5,30] (median 3) vs non-web median 4, with
    # heavy overlap across the whole range and outliers in BOTH classes (one
    # known web has message_count=30, well into what would be "top tertile,
    # penalized" territory). No clean web/non-web separation exists on either
    # field -- kept as reported columns for manual-review context only, not
    # replaced with a different bucketing since none is empirically justified.
    channel_counts = {cid: len(identities[cid]["channels"]) for cid in survivors}
    message_counts = {cid: identities[cid]["message_count"] for cid in survivors}
    print(f"channel_count / message_count: reported only, NOT scored this round (no clean "
          f"web/non-web signal found across the full 189 labeled identities -- see commit message)")

    # --- category_count / address_location_count / off_target_fraction:
    # negative soft signals, thresholds derived from the ACTUAL survivor
    # distribution (top-tertile boundary), not hardcoded ---
    category_counts = {cid: len(evidence[cid]["categories"]) for cid in survivors}
    address_counts = {cid: len(evidence[cid]["address_matches"]) for cid in survivors}
    off_target_fractions = {
        cid: (evidence[cid]["off_target_count"] / evidence[cid]["messages_seen"]
              if evidence[cid]["messages_seen"] else 0.0)
        for cid in survivors
    }
    cat_q1, cat_q2 = tertile_boundaries(category_counts.values())
    addr_q1, addr_q2 = tertile_boundaries(address_counts.values())
    off_q1, off_q2_naive = tertile_boundaries(off_target_fractions.values())
    off_q2 = off_q2_naive

    print(f"\ncategory_count distribution: " +
          ", ".join(f"{k}:{v}" for k, v in sorted(Counter(category_counts.values()).items())))
    print(f"category_count top-tertile boundary (q2) = {cat_q2:.2f} -- "
          f"penalize category_count > {cat_q2:.2f}")

    print(f"\naddress_location_count distribution: " +
          ", ".join(f"{k}:{v}" for k, v in sorted(Counter(address_counts.values()).items())))
    print(f"address_location_count top-tertile boundary (q2) = {addr_q2:.2f} -- "
          f"penalize address_location_count > {addr_q2:.2f}")

    print(f"\noff_target_fraction distribution (rounded to 2dp): " +
          ", ".join(f"{k:.2f}:{v}" for k, v in sorted(Counter(round(x, 2) for x in off_target_fractions.values()).items())))
    print(f"off_target_fraction population-wide top-tertile boundary (q2) = {off_q2_naive:.2f}")

    if off_q2_naive <= 0:
        # Degenerate case: off_target_fraction is heavily zero-inflated
        # (most identities have off_target_fraction==0.0 exactly, unlike the
        # integer counts above which have enough spread near their mode to
        # avoid this). A population-wide tertile boundary of 0 would mean
        # "penalize ANY nonzero value" -- not what "high fraction" should
        # mean. Fall back to the tertile boundary of the NONZERO subset only
        # -- still purely empirical, not a guessed number, just computed over
        # the population where the signal actually varies.
        zero_n = sum(1 for v in off_target_fractions.values() if v == 0)
        nonzero_off = [v for v in off_target_fractions.values() if v > 0]
        print(f"  -- degenerate: {zero_n}/{len(off_target_fractions)} survivors share off_target_fraction="
              f"0.0 exactly, so a population-wide boundary of {off_q2_naive:.2f} would penalize ANY nonzero "
              f"value, not a meaningfully 'high' fraction. Falling back to the tertile boundary of the "
              f"{len(nonzero_off)} NONZERO identities only.")
        if nonzero_off:
            off_q1, off_q2 = tertile_boundaries(nonzero_off)
    print(f"off_target_fraction FINAL threshold used: penalize off_target_fraction > {off_q2:.2f}")

    # --- corp_domain: main-eligibility GATE, not an additive score term ---
    corp_domain_flags = {cid: acr.has_corp_domain(identities[cid]) for cid in survivors}
    print(f"\ncorp_domain=True among survivors: {sum(corp_domain_flags.values())}/{len(survivors)} "
          f"-- these can still be exported/borderline, just never main, regardless of score.")

    # --- score (corp_domain, channel_count, message_count all removed -- see above) ---
    print("\nScore formula (additive integer, deterministic, all inputs SOFT per instruction):")
    print("  employer_count bucket: 2-3 -> +2, 0 -> 0, 1 or 4+ -> -1  (CLAUDE.md hypothesis, unconfirmed)")
    print("  content_volume_flag:   True -> +1, False -> 0")
    print("  has_ref_link:          True -> +1, False -> 0")
    print(f"  category_count:         > {cat_q2:.2f} -> -1, else 0")
    print(f"  address_location_count: > {addr_q2:.2f} -> -1, else 0")
    print(f"  off_target_fraction:    > {off_q2:.2f} -> -1, else 0  (NEW)")
    print("  channel_count/message_count: REMOVED (no clean signal, see above)")
    print("  corp_domain:            REMOVED -- direct main-eligibility gate (see below)")
    print("  has_pvz_mention:        REMOVED -- hard exclude, never reaches scoring (see above)")

    scores = {}
    for cid in survivors:
        ev = evidence[cid]
        s = 0
        s += employer_bucket_score(len(ev["brands"]))
        s += 1 if ev["max_text_len"] >= acr.CONTENT_VOLUME_MIN_CHARS else 0
        s += 1 if ev["has_ref_link"] else 0
        s += -1 if category_counts[cid] > cat_q2 else 0
        s += -1 if address_counts[cid] > addr_q2 else 0
        s += -1 if off_target_fractions[cid] > off_q2 else 0
        scores[cid] = s

    score_values = sorted(scores.values())
    if score_values:
        print(f"\nScore distribution over {len(score_values)} survivors: "
              f"min={score_values[0]} p25={statistics.quantiles(score_values, n=4)[0]:.1f} "
              f"median={statistics.median(score_values):.1f} "
              f"p75={statistics.quantiles(score_values, n=4)[2]:.1f} max={score_values[-1]}")
        hist = Counter(score_values)
        print("Full histogram (score: count): " +
              ", ".join(f"{k}:{v}" for k, v in sorted(hist.items())))

    # --- threshold: score > 0 (net positive evidence), the natural
    # zero-crossing of a signed additive score -- not a percentile chosen to
    # hit a target row count. Unchanged by the corp_domain gate below (that
    # gate only affects WHICH bucket -- main vs borderline -- not whether an
    # identity is exported at all). ---
    THRESHOLD = 0
    exported = {cid: s for cid, s in scores.items() if s > THRESHOLD}
    print(f"\nThreshold: score > {THRESHOLD} (net positive soft evidence). "
          f"Exported: {len(exported)}/{len(survivors)}")

    # corp_domain=True identities are excluded from main REGARDLESS of score
    # (a split-eligibility gate, not a score term) -- they still land in
    # borderline if they clear the export threshold. The main/borderline
    # median split is computed only over corp_domain=False exported
    # identities, since corp_domain=True identities are never main candidates
    # and shouldn't skew where that boundary falls.
    eligible_for_main = {cid: s for cid, s in exported.items() if not corp_domain_flags[cid]}
    forced_borderline = {cid: s for cid, s in exported.items() if corp_domain_flags[cid]}
    print(f"Of the {len(exported)} exported, {len(forced_borderline)} have corp_domain=True and are "
          f"forced to borderline regardless of score; {len(eligible_for_main)} are eligible for main.")

    # main/borderline split by score-threshold PROXIMITY (median split),
    # NOT by content_volume_flag -- see module docstring / report.
    if eligible_for_main:
        split_median = statistics.median(sorted(eligible_for_main.values()))
    else:
        split_median = 0
    main_ids = {cid for cid, s in eligible_for_main.items() if s > split_median}
    borderline_ids = ({cid for cid, s in eligible_for_main.items() if s <= split_median}
                       | set(forced_borderline.keys()))
    print(f"Main/borderline split at median score ({split_median}) of the corp_domain=False "
          f"exported population:")
    print(f"  main (corp_domain=False, score > {split_median}):        {len(main_ids)}")
    print(f"  borderline (corp_domain=True, OR corp_domain=False with score <= {split_median}): "
          f"{len(borderline_ids)}")

    # --- FULL main batch, not samples -- requested explicitly this round ---
    print("\n" + "=" * 78)
    print(f"FULL main batch ({len(main_ids)} rows) -- category + CPA-verified employer "
          f"co-occurrence, corp_domain, category/address counts")
    print("=" * 78)
    all_main_sorted = sorted(main_ids, key=lambda c: -scores[c])
    for i, cid in enumerate(all_main_sorted, 1):
        ev = evidence[cid]
        ex = ev["verified_relevant_examples"][0] if ev["verified_relevant_examples"] else None
        print(f"\n{i}. {cid}  score={scores[cid]}  "
              f"category_count={category_counts[cid]}  address_location_count={address_counts[cid]}  "
              f"off_target_fraction={off_target_fractions[cid]:.2f}  "
              f"(channel_count={channel_counts[cid]} message_count={message_counts[cid]}, reported only)")
        print(f"   brands={sorted(ev['brands'])}  categories={sorted(ev['categories'])}")
        if ev["address_matches"]:
            print(f"   address_matches={sorted(ev['address_matches'])}")
        if ex:
            cats, verified_brands, ex_title = ex
            print(f"   CO-OCCURRED: categories={cats} + verified_brands={verified_brands}")
            print(f"   in message: \"{ex_title[:150]}\"")
        else:
            print(f"   !!! no verified_relevant_examples captured (unexpected -- should be non-empty "
                  f"for anything in main_ids/survivors)")
    print(f"\n(full {len(main_ids)}-row main batch shown above; borderline is {len(borderline_ids)} "
          f"rows, not printed inline)")

    # --- AC-4, now genuinely measurable on the main batch ---
    print("\n" + "=" * 78)
    print("AC-4: Content Sufficiency (now measured -- this population did not exist before Step 4)")
    print("=" * 78)
    if main_ids:
        main_cv_true = sum(1 for cid in main_ids if evidence[cid]["max_text_len"] >= acr.CONTENT_VOLUME_MIN_CHARS)
        pct = 100 * main_cv_true / len(main_ids)
        print(f"candidates_v2_main.xlsx: {main_cv_true}/{len(main_ids)} ({pct:.1f}%) content_volume_flag=True")
        ac4_verdict = "PASS" if pct >= 90 else "FAIL"
        print(f"AC-4 (orientation >=90%): {ac4_verdict} ({pct:.1f}%)")
    else:
        print("main batch is empty -- AC-4 NOT MEASURED")
        ac4_verdict = "NOT MEASURED (empty main batch)"

    # --- write output files ---
    FIELDNAMES = ["вердикт", "canonical_id", "score", "usernames", "phones", "corp_domain",
                  "hr_hint", "has_ref_link", "brands", "employer_count", "multibrand",
                  "categories", "category_count", "has_military_flag", "has_agency_self_id",
                  "has_pvz_mention", "content_volume_flag", "has_verified_relevant_message",
                  "address_location_count", "address_matches", "off_target_fraction",
                  "off_target_examples", "channel_count", "message_count", "first_seen",
                  "last_seen", "sample_titles", "channels", "channel_links"]
    COLUMN_WIDTHS = {
        "вердикт": 14, "canonical_id": 14, "score": 8, "usernames": 45, "phones": 35,
        "corp_domain": 12, "hr_hint": 10, "has_ref_link": 13, "brands": 35,
        "employer_count": 14, "multibrand": 12, "categories": 40, "category_count": 14,
        "has_military_flag": 16, "has_agency_self_id": 18, "has_pvz_mention": 16,
        "content_volume_flag": 18,
        "has_verified_relevant_message": 22, "address_location_count": 20, "address_matches": 60,
        "off_target_fraction": 16, "off_target_examples": 50,
        "channel_count": 14, "message_count": 14, "first_seen": 18, "last_seen": 18,
        "sample_titles": 80, "channels": 55, "channel_links": 60,
    }

    def build_row(cid):
        ident = identities[cid]
        ev = evidence[cid]
        brands = sorted(ev["brands"])
        categories = sorted(ev["categories"])
        channels_list = ident.get("channels") or []
        return {
            "вердикт": "",
            "canonical_id": cid,
            "score": scores[cid],
            "usernames": ", ".join(sorted(c["value"] for c in ident["contacts"] if c["type"] == "username")),
            "phones": ", ".join(sorted(c["value"] for c in ident["contacts"] if c["type"] == "phone")),
            "corp_domain": acr.has_corp_domain(ident),
            "hr_hint": acr.has_hr_hint(ident),
            "has_ref_link": ev["has_ref_link"],
            "brands": ", ".join(brands),
            "employer_count": len(brands),
            "multibrand": len(brands) >= 2,
            "categories": ", ".join(categories),
            "category_count": len(categories),
            "has_military_flag": ev["has_military_flag"],
            "has_agency_self_id": ev["has_agency_self_id"],
            "has_pvz_mention": ev["has_pvz_mention"],
            "content_volume_flag": ev["max_text_len"] >= acr.CONTENT_VOLUME_MIN_CHARS,
            "has_verified_relevant_message": ev["has_verified_relevant_message"],
            "address_location_count": len(ev["address_matches"]),
            "address_matches": ", ".join(f"{name}:{val}" for name, val in sorted(ev["address_matches"])),
            "off_target_fraction": round(ev["off_target_count"] / ev["messages_seen"], 3) if ev["messages_seen"] else 0.0,
            "off_target_examples": " | ".join(ev["off_target_examples"]),
            "channel_count": len(channels_list),
            "message_count": ident["message_count"],
            "first_seen": acr.parse_dt(ident["first_seen"]),
            "last_seen": acr.parse_dt(ident["last_seen"]),
            "sample_titles": " | ".join(ev["titles"]),
            "channels": ", ".join(channels_list),
            "channel_links": " ".join(f"https://t.me/s/{ch}" for ch in channels_list),
        }

    def write_export(cids, path):
        rows = sorted((build_row(cid) for cid in cids), key=lambda r: -r["score"])
        wb = Workbook()
        ws = wb.active
        ws.title = "candidates"
        ws.append(FIELDNAMES)
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial")
        for r in rows:
            ws.append([r[k] for k in FIELDNAMES])
        dv = DataValidation(type="list", formula1='"веб,не_веб,агентство,не_уверен"', allow_blank=True)
        dv.errorTitle = "Недопустимое значение"
        dv.error = "Выбери значение из списка: веб, не_веб, агентство, не_уверен"
        ws.add_data_validation(dv)
        dv.add(f"A2:A{max(len(rows) + 1, 2)}")
        for i, header in enumerate(FIELDNAMES, start=1):
            ws.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTHS[header]
        ws.freeze_panes = "A2"
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return len(rows)

    n_main = write_export(main_ids, OUTPUT_MAIN)
    n_border = write_export(borderline_ids, OUTPUT_BORDERLINE)
    print(f"\nWrote {OUTPUT_MAIN.relative_to(BASE_DIR)}: {n_main} rows")
    print(f"Wrote {OUTPUT_BORDERLINE.relative_to(BASE_DIR)}: {n_border} rows")
    print(f"\nAC-4 verdict: {ac4_verdict}")


if __name__ == "__main__":
    main()

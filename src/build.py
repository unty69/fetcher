"""
build.py — финальная сборка: cache/dump.jsonl -> output/webs.xlsx.

Квалифицирует сообщения (ref_domains / erid / tg://resolve), резолвит и
идентифицирует квалифицирующие http(s)-ссылки через существующие
resolve.py/identify.py (логика там не меняется, только вызывается),
группирует контакты и схлопывает дубли одного веба под разными контактами.
"""

import argparse
import csv
import json
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl

# Windows-консоль по умолчанию не всегда в UTF-8 — без этого print()
# с кириллицей может упасть с UnicodeEncodeError.
sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import identify  # noqa: E402
import resolve  # noqa: E402
from config import CPA_URL_MARKERS, HOTLINE_PHONE_RE, MFO_DOMAINS, MFO_ID_PARAMS, MILITARY_KEYWORDS  # noqa: E402

REF_DOMAINS_PATH = BASE_DIR / "input" / "ref_domains.txt"
DUMP_PATH = BASE_DIR / "cache" / "dump.jsonl"
OUTPUT_PATH = BASE_DIR / "output" / "webs.xlsx"
VERDICTS_PATH = BASE_DIR / "output" / "webs_verdict.csv"
VERDICTS_HEADER = ["web_key", "вердикт"]

TG_RESOLVE_PREFIX = "tg://resolve?domain="


# --- квалификация сообщений/URL ------------------------------------------


def load_ref_domains() -> set:
    domains = set()
    with open(REF_DOMAINS_PATH, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            domains.add(line.lower())
    return domains


def _with_scheme(url: str) -> str:
    """urlparse не находит netloc без схемы (напр. "clck.ru/xyz") -- без
    "://" в url подставляем "https://" перед парсингом."""
    return url if "://" in url else "https://" + url


def matches_ref_domain(url: str, ref_domains: set) -> bool:
    netloc = urlparse(_with_scheme(url)).netloc.lower()
    return any(domain in netloc for domain in ref_domains)


def has_erid(url: str) -> bool:
    return "erid=" in url or "erid%3D" in url


def has_cpa_marker(url: str) -> bool:
    return any(marker in url for marker in CPA_URL_MARKERS)


def is_tg_resolve(url: str) -> bool:
    return url.startswith(TG_RESOLVE_PREFIX)


def extract_tg_domain(url: str):
    values = parse_qs(urlparse(url).query).get("domain")
    return values[0] if values else None


def is_military_content(text: str) -> bool:
    lowered = text.lower()
    return any(keyword in lowered for keyword in MILITARY_KEYWORDS)


# --- нормализация контактов -----------------------------------------------


def normalize_phone(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    if len(digits) == 11 and digits[0] in ("7", "8"):
        return "7" + digits[1:]
    if len(digits) == 10:
        return "7" + digits
    return digits


def normalize_contact(contact_type: str, value: str) -> str:
    if contact_type == "phone":
        return normalize_phone(value)
    if contact_type == "email":
        return value.lower()
    if contact_type == "username":
        value = value.lower()
        return value[1:] if value.startswith("@") else value
    return value


HOTLINE_PHONE_REGEX = re.compile(HOTLINE_PHONE_RE)


def is_hotline_phone(normalized: str) -> bool:
    """normalized -- 11-значный номер после normalize_phone (всегда с '7' в
    начале). normalize_phone стирает разницу между исходными 7... и 8..., так
    что горячую линию 8-800-... нужно проверять на варианте с '8' обратно."""
    variant = "8" + normalized[1:]
    return bool(HOTLINE_PHONE_REGEX.match(variant))


def filter_hotline_contacts(contacts: list):
    """Убирает контакты-горячие линии (шаг 4 правки). Возвращает (filtered,
    disqualified) -- disqualified=True, если контакты были, а после фильтра
    не осталось ни одного (сообщение из-за этого перестаёт квалифицироваться)."""
    had_contacts = len(contacts) > 0
    filtered = [
        c for c in contacts
        if not (c["type"] == "phone" and is_hotline_phone(normalize_phone(c["value"])))
    ]
    disqualified = had_contacts and not filtered
    return filtered, disqualified


# --- проход по дампу: квалификация + разделение tg:// / http(ы) ----------


def load_dump_and_classify(dump_paths: list, ref_domains: set):
    """dump_paths -- список путей; читаются и объединяются как один поток
    (просто конкатенация JSONL-строк, порядок между файлами не важен,
    каждая строка независима). Возвращает (число квалифицирующих сообщений,
    military_excluded, hotline_excluded, tg_repeats, web_messages,
    множество уникальных http(s)-URL для резолва)."""
    qualifying_count = 0
    military_excluded = 0
    hotline_excluded = 0
    tg_repeats = defaultdict(lambda: {"channels": set(), "count": 0})
    web_messages = []
    unique_resolve_urls = set()

    for dump_path in dump_paths:
        with open(dump_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                msg = json.loads(line)

                if is_military_content(msg["text"]):
                    military_excluded += 1
                    continue

                urls = [u["url"] for u in msg["urls"]]

                tg_urls = [u for u in urls if is_tg_resolve(u)]
                web_urls = [
                    u for u in urls
                    if not is_tg_resolve(u)
                    and (matches_ref_domain(u, ref_domains) or has_erid(u) or has_cpa_marker(u))
                ]

                if not tg_urls and not web_urls:
                    continue  # ни одного квалифицирующего условия -- пропуск целиком

                # web_urls квалифицирует сообщение только если после фильтра
                # горячих линий у него остался хотя бы один контакт (или
                # контактов не было вообще -- тогда фильтр горячих линий тут
                # ни при чём).
                web_qualifies = False
                filtered_contacts = []
                if web_urls:
                    filtered_contacts, disqualified = filter_hotline_contacts(msg["contacts"])
                    if disqualified:
                        hotline_excluded += 1
                    else:
                        web_qualifies = True

                if not tg_urls and not web_qualifies:
                    continue  # было только web_urls, и фильтр горячих линий его снял

                qualifying_count += 1

                for u in tg_urls:
                    domain = extract_tg_domain(u)
                    if domain:
                        tg_repeats[domain]["channels"].add(msg["channel"])
                        tg_repeats[domain]["count"] += 1

                if web_qualifies:
                    web_messages.append({
                        "channel": msg["channel"],
                        "message_id": msg["message_id"],
                        "date": msg["date"],
                        "text": msg["text"],
                        "contacts": filtered_contacts,
                        "urls": web_urls,
                    })
                    unique_resolve_urls.update(web_urls)

    return (qualifying_count, military_excluded, hotline_excluded, tg_repeats,
            web_messages, unique_resolve_urls)


# --- группировка контактов -------------------------------------------------


MFO_DOMAINS = {d.lower() for d in MFO_DOMAINS}


def is_non_hr_message(urls: list, resolved: dict) -> bool:
    """True, если хотя бы один url сообщения резолвится в запись из
    MFO_DOMAINS -- подстрокой в ПОЛНОМ final_url (не только в domain_final):
    домены-хабы вроде bankpro.su отдают и займы (/category/14-mikrokrediti),
    и вакансии (/vacancy) с одного домена -- матч по одному domain_final
    увёл бы реальные вакансии в "МФО-вебы". Для обычных МФО-доменов без "/"
    в записи поведение не меняется -- домен всё равно подстрока final_url.
    Технически квалифицируется, но маршрутизируется в отдельный лист
    "МФО-вебы", а не "Вебы" (см. main())."""
    for u in urls:
        final_url = resolved[u].get("final_url") or u
        if any(d in final_url for d in MFO_DOMAINS):
            return True
    return False


def message_networks_and_ids(urls: list, identified: dict):
    networks = set()
    identity_ids = set()
    for u in urls:
        res = identified[u]
        if res["network"]:
            networks.add(res["network"])
        if res["has_personal_id"]:
            for values in res["identity_params"].values():
                for v in values:
                    identity_ids.add((res["network"], v))
    return networks, identity_ids


def aggregate_contacts(web_messages: list, identified: dict) -> dict:
    """Агрегирует ЛЮБОЙ переданный список сообщений по контакту -- маршрут
    (HR/МФО) в эту функцию не встроен, список уже отфильтрован вызывающим
    кодом (см. main()).

    Сообщение без контакта, но с identity_ids (has_personal_id=True по
    NETWORK_RULES), раньше терялось молча: цикл "for c in wm['contacts']"
    просто не выполнялся, хотя identity уже была посчитана строкой выше.
    Теперь для такого сообщения синтезируется одна запись с
    contact_type="identity_only". merge_by_identity() ниже схлопывает записи
    по общему identity_id, так что синтетическая запись сама объединится с
    реальным контактом того же веба, если такой контакт найдётся в другом
    сообщении -- дублирования строк не создаёт."""
    contacts_agg = {}

    for wm in web_messages:
        networks, identity_ids = message_networks_and_ids(wm["urls"], identified)
        dt = datetime.fromisoformat(wm["date"])

        contacts = wm["contacts"]
        if not contacts:
            if not identity_ids:
                continue  # ни контакта, ни identity -- строку сделать не из чего
            id_repr = "|".join(sorted(f"{net}:{val}" for net, val in identity_ids))
            contacts = [{"type": "identity_only", "value": id_repr}]

        for c in contacts:
            normalized = c["value"] if c["type"] == "identity_only" else normalize_contact(c["type"], c["value"])
            key = (c["type"], normalized)
            agg = contacts_agg.setdefault(key, {
                "contact_type": c["type"],
                "channels": set(),
                "networks": set(),
                "identity_ids": set(),
                "message_ids": set(),
                "first_seen": None,
                "last_seen": None,
                "example_text": None,
            })

            agg["channels"].add(wm["channel"])
            agg["networks"] |= networks
            agg["identity_ids"] |= identity_ids
            agg["message_ids"].add((wm["channel"], wm["message_id"]))

            if agg["first_seen"] is None or dt < agg["first_seen"]:
                agg["first_seen"] = dt
                agg["example_text"] = wm["text"][:300]
            if agg["last_seen"] is None or dt > agg["last_seen"]:
                agg["last_seen"] = dt

    return contacts_agg


def _find(parent: dict, x):
    while parent[x] != x:
        parent[x] = parent[parent[x]]
        x = parent[x]
    return x


def _union(parent: dict, a, b):
    ra, rb = _find(parent, a), _find(parent, b)
    if ra != rb:
        parent[ra] = rb


def _merge_group(contacts_agg: dict, group_keys: list) -> dict:
    contact_values = []
    contact_types = []
    channels = set()
    networks = set()
    identity_ids = set()
    message_ids = set()
    first_seen = None
    last_seen = None
    example_text = None

    for key in group_keys:
        agg = contacts_agg[key]
        contact_values.append(key[1])
        if agg["contact_type"] not in contact_types:
            contact_types.append(agg["contact_type"])
        channels |= agg["channels"]
        networks |= agg["networks"]
        identity_ids |= agg["identity_ids"]
        message_ids |= agg["message_ids"]
        if agg["first_seen"] is not None and (first_seen is None or agg["first_seen"] < first_seen):
            first_seen = agg["first_seen"]
            example_text = agg["example_text"]
        if agg["last_seen"] is not None and (last_seen is None or agg["last_seen"] > last_seen):
            last_seen = agg["last_seen"]

    if identity_ids:
        web_key = "id:" + "|".join(sorted(f"{net}:{val}" for net, val in identity_ids))
    else:
        web_key = "contact:" + "|".join(sorted(contact_values))

    return {
        "web_key": web_key,
        "contacts": ", ".join(contact_values),
        "contact_type": ", ".join(contact_types),
        "networks": ", ".join(sorted(networks)),
        "channels": ", ".join(sorted(channels)),
        "message_count": len(message_ids),
        "first_seen": first_seen,
        "last_seen": last_seen,
        "needs_review": len(networks) == 0,
        "example_text": example_text,
    }


def merge_by_identity(contacts_agg: dict) -> list:
    """Схлопывает контакты, у которых пересекаются identity_ids (шаг 8):
    общая (network, id) пара -- один веб под разными контактами."""
    keys = list(contacts_agg.keys())
    parent = {k: k for k in keys}

    id_to_keys = defaultdict(list)
    for key, agg in contacts_agg.items():
        for iid in agg["identity_ids"]:
            id_to_keys[iid].append(key)

    for keys_sharing in id_to_keys.values():
        for k in keys_sharing[1:]:
            _union(parent, keys_sharing[0], k)

    groups = defaultdict(list)
    for key in keys:
        groups[_find(parent, key)].append(key)

    return [_merge_group(contacts_agg, group_keys) for group_keys in groups.values()]


# --- группировка МФО-вебов: по ID в итоговом URL, не по контакту -----------


def extract_mfo_id(final_url: str):
    """Займовые лендинги не несут контакта в тексте -- воронка целиком через
    клик по ссылке. Идентичность веба тут не в контакте, а в параметрах
    итогового URL (MFO_ID_PARAMS). Возвращает значение первого найденного
    (по порядку MFO_ID_PARAMS) параметра, иначе None."""
    query = parse_qs(urlparse(final_url).query)
    for param in MFO_ID_PARAMS:
        if param in query and query[param]:
            return query[param][0]
    return None


def aggregate_mfo(mfo_messages: list, resolved: dict, identified: dict) -> dict:
    """Параллель aggregate_contacts(), но группирует не по контакту (МФО-
    сообщения его обычно не несут -- см. extract_mfo_id), а по MFO-ID.
    Дедуп по identity (merge_by_identity) тут не нужен -- группировка по ID
    уже и есть дедуп, отдельного объединения не требуется."""
    mfo_agg = {}

    for wm in mfo_messages:
        mfo_id = None
        for u in wm["urls"]:
            mfo_id = extract_mfo_id(resolved[u]["final_url"])
            if mfo_id is not None:
                break

        key = ("mfo_id", mfo_id) if mfo_id is not None else ("channel_only", wm["channel"])
        dt = datetime.fromisoformat(wm["date"])

        agg = mfo_agg.setdefault(key, {
            "domains": set(),
            "channels": set(),
            "message_ids": set(),
            "first_seen": None,
            "last_seen": None,
            "example_text": None,
        })

        for u in wm["urls"]:
            agg["domains"].add(identified[u]["domain_final"])
        agg["channels"].add(wm["channel"])
        agg["message_ids"].add((wm["channel"], wm["message_id"]))

        if agg["first_seen"] is None or dt < agg["first_seen"]:
            agg["first_seen"] = dt
            agg["example_text"] = wm["text"][:300]
        if agg["last_seen"] is None or dt > agg["last_seen"]:
            agg["last_seen"] = dt

    return mfo_agg


def mfo_rows_from_agg(mfo_agg: dict) -> list:
    """key = ("mfo_id", id) -> web_key "mfo_id:<id>", needs_id=False;
    key = ("channel_only", channel) -> web_key "mfo_channel:<channel>",
    needs_id=True (fallback -- сообщение не потеряно, но увереность ниже)."""
    rows = []
    for (kind, value), agg in mfo_agg.items():
        if kind == "mfo_id":
            web_key = "mfo_id:" + value
            affiliate_id = value
            needs_id = False
        else:
            web_key = "mfo_channel:" + value
            affiliate_id = ""
            needs_id = True

        rows.append({
            "web_key": web_key,
            "affiliate_id": affiliate_id,
            "domains": ", ".join(sorted(agg["domains"])),
            "channels": ", ".join(sorted(agg["channels"])),
            "message_count": len(agg["message_ids"]),
            "first_seen": agg["first_seen"],
            "last_seen": agg["last_seen"],
            "needs_id": needs_id,
            "example_text": agg["example_text"],
        })
    return rows


# --- ручная разметка между прогонами ---------------------------------------


def load_verdicts(path: Path) -> dict:
    """{web_key: вердикт} из output/webs_verdict.csv, если файл существует
    (человек мог дописать туда разметку руками после предыдущего прогона).
    Файла нет -- {} (первый прогон, разметки ещё не было)."""
    if not path.exists():
        return {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        return {
            row["web_key"]: row.get("вердикт") or ""
            for row in csv.DictReader(f)
            if row.get("web_key")
        }


def save_verdicts(verdicts: dict, path: Path) -> None:
    """Пишет {web_key: вердикт} в csv -- тот же формат, что читает
    load_verdicts(). Отсортировано по web_key для стабильного вывода между
    прогонами (удобно смотреть diff файла)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(VERDICTS_HEADER)
        for web_key, verdict in sorted(verdicts.items()):
            writer.writerow([web_key, verdict])


def ensure_verdicts_file(path: Path) -> None:
    """Создаёт пустой output/webs_verdict.csv с одним заголовком, если файла
    ещё нет -- чтобы было видно, куда дописывать ручную разметку."""
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerow(VERDICTS_HEADER)


def _harvest_verdicts_from_xlsx(xlsx_path: Path, csv_path: Path) -> None:
    """Логика harvest -- БЕЗ защиты от исключений. Не вызывать из main()
    напрямую (см. harvest_verdicts_from_xlsx() ниже) -- эта версия для
    прямой диагностики (разовый скрипт), где нужен настоящий traceback, а
    не проглоченное сообщение. Идёт по обоим листам ("Вебы" и "МФО-вебы") --
    у обоих есть web_key и мой_вердикт, хоть остальная схема и разная."""
    if not xlsx_path.exists():
        return

    harvested = {}
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for sheet_name in ("Вебы", "МФО-вебы"):
            if sheet_name not in wb.sheetnames:
                continue
            rows = wb[sheet_name].iter_rows(values_only=True)
            header = next(rows, None)
            if header is None or "web_key" not in header or "мой_вердикт" not in header:
                continue
            web_key_col = header.index("web_key")
            verdict_col = header.index("мой_вердикт")
            last_needed_col = max(web_key_col, verdict_col)

            for row in rows:
                # openpyxl (read_only) может отдать укороченный tuple для
                # строки с пустыми хвостовыми ячейками -- не считаем это
                # поводом ронять весь харвест, просто пропускаем строку.
                if len(row) <= last_needed_col:
                    continue
                web_key, verdict = row[web_key_col], row[verdict_col]
                if web_key and verdict and str(verdict).strip():
                    harvested[str(web_key).strip()] = str(verdict).strip()
    finally:
        wb.close()

    if not harvested:
        return

    verdicts = load_verdicts(csv_path)
    verdicts.update(harvested)
    save_verdicts(verdicts, csv_path)
    print(f"harvest_verdicts_from_xlsx: подобрано {len(harvested)} вердиктов из {xlsx_path.name}")


def harvest_verdicts_from_xlsx(xlsx_path: Path, csv_path: Path) -> None:
    """Перед пересборкой -- подобрать ручные пометки "мой_вердикт", если
    человек вписал их прямо в output/webs.xlsx (в Excel), а не в csv, и
    перенести в webs_verdict.csv ДО того, как xlsx будет перезаписан заново
    (иначе собирать уже будет нечего -- файл станет новым). xlsx -- внешний
    файл, который правит человек в Excel, поэтому чтение обёрнуто: не даём
    одной сломанной книге (открыта в Excel, битый файл и т.п.) остановить
    всю пересборку. Это ТОЛЬКО защита от чтения xlsx -- отдельно от неё в
    main() есть safety-net на случай, если она сама тут не сработает (см.
    backup_webs_xlsx())."""
    try:
        _harvest_verdicts_from_xlsx(xlsx_path, csv_path)
    except Exception as e:
        print(f"harvest_verdicts_from_xlsx: не удалось прочитать {xlsx_path} ({e!r}), пропускаю")


def backup_webs_xlsx(path: Path) -> None:
    """Резервная копия path -> path с добавленным ".bak" ПЕРЕД перезаписью
    (main() зовёт это прямо перед write_workbook()) -- независимо от того,
    прошёл ли harvest_verdicts_from_xlsx() успешно. Второй рубеж защиты
    поверх try/except в harvest: если тот всё-таки где-то незаметно не
    подхватит вердикты, .bak -- это версия файла ДО перезаписи, откуда их
    можно вытащить руками, а не потерять без возврата. Хранится одно
    поколение назад, не архив -- .bak перезаписывается каждый прогон."""
    if not path.exists():
        return
    backup_path = path.with_name(path.name + ".bak")
    shutil.copy2(path, backup_path)


# --- вывод -----------------------------------------------------------------


WEBS_HEADER = ["web_key", "contact(s)", "contact_type", "networks", "channels",
               "message_count", "first_seen", "last_seen", "needs_review", "example_text",
               "мой_вердикт"]


def _write_webs_sheet(ws, rows: list, verdicts: dict) -> None:
    """Лист "Вебы" (контакт-центричная схема). "МФО-вебы" -- отдельная схема,
    см. _write_mfo_sheet(); оба листа читают один и тот же verdicts
    (webs_verdict.csv один на оба листа, а не два разных файла)."""
    ws.append(WEBS_HEADER)
    for row in rows:
        ws.append([
            row["web_key"],
            row["contacts"],
            row["contact_type"],
            row["networks"],
            row["channels"],
            row["message_count"],
            row["first_seen"].replace(tzinfo=None) if row["first_seen"] else None,
            row["last_seen"].replace(tzinfo=None) if row["last_seen"] else None,
            row["needs_review"],
            row["example_text"],
            verdicts.get(row["web_key"], ""),
        ])


MFO_HEADER = ["web_key", "affiliate_id", "домены", "channels", "message_count",
              "first_seen", "last_seen", "needs_id", "example_text", "мой_вердикт"]


def _write_mfo_sheet(ws, rows: list, verdicts: dict) -> None:
    """Схема "МФО-вебы" -- осознанно не копия "Вебы": нет contact(s)/
    contact_type/networks (МФО-лендинги не несут контакта и не проходят
    NETWORK_RULES), вместо этого affiliate_id/домены/needs_id."""
    ws.append(MFO_HEADER)
    for row in rows:
        ws.append([
            row["web_key"],
            row["affiliate_id"],
            row["domains"],
            row["channels"],
            row["message_count"],
            row["first_seen"].replace(tzinfo=None) if row["first_seen"] else None,
            row["last_seen"].replace(tzinfo=None) if row["last_seen"] else None,
            row["needs_id"],
            row["example_text"],
            verdicts.get(row["web_key"], ""),
        ])


def write_workbook(webs_rows: list, mfo_rows: list, tg_repeat_rows: list, verdicts: dict) -> None:
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Вебы"
    _write_webs_sheet(ws1, webs_rows, verdicts)

    ws2 = wb.create_sheet("Повторы между каналами")
    ws2.append(["domain", "channels", "message_count"])
    for row in tg_repeat_rows:
        ws2.append([row["domain"], row["channels"], row["message_count"]])

    ws3 = wb.create_sheet("МФО-вебы")
    _write_mfo_sheet(ws3, mfo_rows, verdicts)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)


def build_arg_parser() -> argparse.ArgumentParser:
    """--dump-path необязательный и повторяемый: без него поведение не
    меняется вообще (default=None -> main() берёт [DUMP_PATH], как раньше).
    С несколькими -- источники читаются и объединяются в один поток ДО
    квалификации; сама квалификация/резолв/группировка их не различают."""
    parser = argparse.ArgumentParser(description="Собрать output/webs.xlsx из одного или нескольких JSONL-дампов.")
    parser.add_argument("--dump-path", action="append", default=None,
                         help=f"JSONL-дамп для чтения, можно указать несколько раз "
                              f"(по умолчанию {DUMP_PATH.relative_to(BASE_DIR)})")
    return parser


def display_path(p: Path) -> str:
    try:
        return str(p.relative_to(BASE_DIR))
    except ValueError:
        return str(p)


def resolve_dump_paths(values) -> list:
    if values is None:
        return [DUMP_PATH]
    resolved = []
    for v in values:
        p = Path(v)
        resolved.append(p if p.is_absolute() else BASE_DIR / p)
    return resolved


def count_lines(path: Path) -> int:
    count = 0
    with open(path, encoding="utf-8") as f:
        for line in f:
            if line.strip():
                count += 1
    return count


def main():
    harvest_verdicts_from_xlsx(OUTPUT_PATH, VERDICTS_PATH)

    args = build_arg_parser().parse_args()
    dump_paths = resolve_dump_paths(args.dump_path)

    per_file_counts = [(p, count_lines(p)) for p in dump_paths]
    for p, n in per_file_counts:
        print(f"{display_path(p)}: {n} сообщений")
    print(f"Всего сообщений (все файлы): {sum(n for _, n in per_file_counts)}")

    ref_domains = load_ref_domains()

    (qualifying_count, military_excluded, hotline_excluded, tg_repeats,
     web_messages, unique_resolve_urls) = load_dump_and_classify(dump_paths, ref_domains)

    try:
        with open(resolve.CACHE_FILE, encoding="utf-8") as f:
            existing_cache = json.load(f)
    except FileNotFoundError:
        existing_cache = {}
    new_count = sum(1 for u in unique_resolve_urls if u not in existing_cache)
    print(f"Уникальных URL для резолва: {len(unique_resolve_urls)} "
          f"(в кэше: {len(unique_resolve_urls) - new_count}, новых: {new_count})")

    resolved = resolve.resolve_batch(list(unique_resolve_urls))
    identified = {u: identify.identify(u, resolved[u]) for u in unique_resolve_urls}

    hr_messages = [wm for wm in web_messages if not is_non_hr_message(wm["urls"], resolved)]
    mfo_messages = [wm for wm in web_messages if is_non_hr_message(wm["urls"], resolved)]

    contacts_agg = aggregate_contacts(hr_messages, identified)
    final_rows = merge_by_identity(contacts_agg)
    final_rows.sort(key=lambda r: r["needs_review"])

    mfo_agg = aggregate_mfo(mfo_messages, resolved, identified)
    final_rows_mfo = mfo_rows_from_agg(mfo_agg)
    final_rows_mfo.sort(key=lambda r: r["needs_id"])

    tg_repeat_rows = [
        {"domain": domain, "channels": ", ".join(sorted(info["channels"])), "message_count": info["count"]}
        for domain, info in tg_repeats.items()
        if len(info["channels"]) >= 2
    ]

    ensure_verdicts_file(VERDICTS_PATH)
    verdicts = load_verdicts(VERDICTS_PATH)

    backup_webs_xlsx(OUTPUT_PATH)
    write_workbook(final_rows, final_rows_mfo, tg_repeat_rows, verdicts)

    print(f"Квалифицирующих сообщений: {qualifying_count} (было 3027)")
    print(f"military_excluded: {military_excluded}")
    print(f"hotline_excluded: {hotline_excluded}")
    print(f"Сообщений в HR-вебы: {len(hr_messages)}")
    print(f"Сообщений в МФО-вебы: {len(mfo_messages)}")
    print(f"Итоговых строк в 'Вебы': {len(final_rows)}")
    print(f"needs_review=True: {sum(1 for r in final_rows if r['needs_review'])}")
    print(f"Итоговых строк в 'МФО-вебы': {len(final_rows_mfo)}")
    print(f"needs_id=True (МФО, fallback по каналу): {sum(1 for r in final_rows_mfo if r['needs_id'])}")

    mfo_ids = {value for kind, value in mfo_agg if kind == "mfo_id"}
    known_spam_cluster = {"250236", "153920"}
    print(f"Уникальных MFO-ID: {len(mfo_ids)}")
    print(f"Из них вне кластера 250236/153920: {len(mfo_ids - known_spam_cluster)}")


if __name__ == "__main__":
    main()

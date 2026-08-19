"""
search_employers.py — ищет вакансии, где ОДНОВРЕМЕННО: (1) заголовок
сообщения соответствует целевой категории uHead, и (2) в тексте
упомянут отслеживаемый работодатель без referen-контекста.

Правила JOB_CATEGORY_KEYWORDS используют \b слева от каждого корня --
без этого "водитель" ловил "руководитель" как подстроку (реальный
баг, найден на разметке: оба "Руководитель по продажам"/"Руководитель
отдела маркетинга" ошибочно попадали в категорию "Водители"). Справа
границу не ставим -- иначе теряются словоформы "водителя"/"водителей".

likely_corp_hr (домен) и has_promo_tone -- МЯГКИЕ сигналы сортировки,
не жёсткие фильтры. Проверено на 86 реально размеченных строках:
жёсткое исключение по "своему домену" срезало бы единственный
подтверждённый веб (tkadysheva@happy-phone.ru, "Выездной представитель
МТС банка"), жёсткое исключение по отсутствию промо-тона срезало бы
ещё 2 подтверждённых веба ("Выездной представитель МТС банка",
"Водитель-курьер (Доставка ценностей)" -- оба написаны сухо, без
эмодзи/CTA, но реальные веб). Поэтому оба сигнала только двигают
строку вниз списка, никогда не убирают её совсем.

has_promo_tone проверено на той же разметке: ловит 63 из 69 "не веб"
(91%), 5 ложных срабатываний на "веб" (ровно те два случая выше).
has_address подмножество has_promo_tone почти полностью (совпадающие
числа при тесте), оставлен отдельной колонкой для прозрачности, но
не добавляет отдельного веса в скор.

GENERIC_PLATFORM_DOMAINS/EMPLOYER_COUNT_NOISE_THRESHOLD/
already_visible -- по-прежнему ЖЁСТКИЕ исключения (не менялись,
были провалидированы раньше на двух подтверждённых случаях jobers.ru
и +74951832415).
"""
import argparse, csv, json, re, sys
from pathlib import Path
from collections import defaultdict

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import build  # noqa: E402

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

DUMP_PATHS = [BASE_DIR / "cache" / n for n in
              ["dump.jsonl", "dump_tier2.jsonl", "dump_tier3.jsonl", "dump_depth_probe.jsonl"]]
OUTPUT_JSON = BASE_DIR / "cache" / "employer_mentions.json"
OUTPUT_XLSX = BASE_DIR / "output" / "employer_mentions_review.xlsx"

GENERIC_PLATFORM_DOMAINS = ("jobers.ru", "hh.ru", "headhunter.ru")
EMPLOYER_COUNT_NOISE_THRESHOLD = 3

JOB_CATEGORY_KEYWORDS = {
    "Курьеры и доставка": r"\bкурьер|\bдоставщик",
    "Склад": r"\bсборщик|\bкомплектовщик|\bупаковщик|\bсортировщик|\bгрузчик|\bкладовщик|\bмаркировщик|\bоператор\s+(штабелера|погрузчика)|\bприёмщик\s+товара|\bприемщик\s+товара|\b(?:работник|сотрудник)\s+склада",
    "Водители": r"\bводитель",
    "Розничная торговля": r"\bпродавец|\bкассир|\bмерчандайзер|\bадминистратор\s+(магазина|торгового\s+зала)|\bсотрудник\s+магазина",
    "Удалённая работа": r"\bоператор\s+(call|колл|контактного)[\s-]*центра|\bтелемаркетолог|\bмодератор|\bдиспетчер",
    "Общепит": r"\bповар|\bофициант|\bбариста|\bбармен|\bпиццамейкер|\bмойщик\s+посуды|\bхостес|\bсотрудник\s+ресторана",
    "Квалифицированные рабочие": r"\bстроител|\bремонтник|\bэлектрик|\bсантехник|\bмонтажник",
    "[КАНДИДАТ] Клининг": r"\bклинер|\bуборщи[кц]",
    "[КАНДИДАТ] Разъездные специалисты": r"\bвыездн\w+\s+(представител|специалист)|\bмобильн\w+\s+банкир",
}

EMPLOYERS = {
    "4 Лапы": r"4\s*Лапы|Четыре\s*Лапы", "Burger King": r"Burger\s*King", "Efin": r"\bEfin\b",
    "Ozon": r"\bOzon\b|\bОзон\b",
    "Qlean": r"\bQlean\b", "ROSTIC'S": r"ROSTIC'?S|Ростикс", "VOXYS": r"\bVOXYS\b",
    "Ventra": r"\bVentra\b|\bВентра\b", "X5 Доставка": r"X5\s*Доставка|X5\s*Group",
    "lamoda": r"\blamoda\b|Ламода", "Альфа-Банк": r"Альфа-?Банк", "ВкусВилл": r"ВкусВилл",
    "Газпромнефть": r"Газпромнефть", "Дикси": r"\bДикси\b", "Домовёнок": r"Домовён[оё]к",
    "Купер": r"\bКупер\b|Сбермаркет|СберМаркет|SberMarket", "МТС": r"\bМТС\b",
    "Магнит Заряд": r"Магнит\s*[\"«]?Заряд", "Магнит доставка": r"Магнит\s*доставка",
    "ОНЕКТА": r"\bОНЕКТА\b", "Т-Банк": r"Т-?Банк", "Тетрика": r"Тетрика",
    "Яндекс Еда": r"Яндекс\s*Еда", "Яндекс Лавка": r"Яндекс\s*Лавка", "Яндекс Маркет": r"Яндекс\s*Маркет",
    "Восток Запад Логистика": r"Восток\s*Запад\s*Логистика", "Клин Лабс": r"Клин\s*Лабс",
    "Достависта": r"Достависта", "ВТБ": r"\bВТБ\b",
}

CORP_DOMAINS = {
    "Ozon": ("ozon.ru",), "Т-Банк": ("tbank.ru", "tinkoff.ru"),
    "HeadHunter": ("hh.ru", "headhunter.ru"), "Яндекс Маркет": ("yandex.ru", "yandex-team.ru"),
    "Яндекс Еда": ("yandex.ru", "yandex-team.ru"), "Яндекс Лавка": ("yandex.ru", "yandex-team.ru"),
    "ВкусВилл": ("vkusvill.ru",), "Дикси": ("dixy.ru",), "МТС": ("mts.ru", "mtsbank.ru", "mtsretail.ru"),
    "lamoda": ("lamoda.ru",), "Альфа-Банк": ("alfabank.ru",), "Burger King": ("burgerking.ru", "rest.group"),
    "VOXYS": ("voxys.ru",), "ROSTIC'S": ("rostics.ru", "cityrst.ru", "uni.rest"),
    "Газпромнефть": ("gazprom-neft.ru",), "Ventra": ("ventra.ru", "ventra.biz"),
    "Купер": ("kuper.ru", "sbermarket.ru"), "X5 Доставка": ("x5.ru", "x5group.ru"),
    "Qlean": ("qlean.ru",), "4 Лапы": ("4lapy.ru",),
    "Достависта": ("dostavista.ru",), "ВТБ": ("vtb.ru",),
}
PERSONAL_EMAIL_PROVIDERS = ("mail.ru", "yandex.ru", "yandex.com", "gmail.com", "bk.ru",
                             "inbox.ru", "list.ru", "rambler.ru", "icloud.com", "outlook.com",
                             "hotmail.com", "ya.ru")

REFERENCE_TRIGGERS = re.compile(
    r'(плюсом|приветствуется|предпочтительно|желательно|опыт\s+(?:работы\s+)?в|'
    r'стаж\s+(?:работы\s+)?в|рассматрива\w*\s+(?:такие\s+)?запрос|ищем\s+(?:также\s+)?по\s+запрос)',
    re.IGNORECASE,
)

WEB_SIGNAL_RE = re.compile(r'партн[её]р|partner|команд', re.IGNORECASE)
COURIER_EMPLOYERS = {"Яндекс Еда", "Яндекс Лавка", "Купер", "Достависта",
                      "Магнит доставка", "X5 Доставка", "Самокат"}


def compute_web_score(employers: list, titles: list, promo: bool) -> int:
    score = 0
    joined = " | ".join(titles)
    if WEB_SIGNAL_RE.search(joined):
        score += 2
    if any(e in COURIER_EMPLOYERS for e in employers):
        score += 1
    if promo:
        score += 1
    return score


ADDRESS_RE = re.compile(
    r'\((?:[^)]*\b(?:м\.?|ул\.?|пр-?кт|пр\.|пер\.?|наб\.?|г\.|тц|шоссе|ш\.|район|д\.?\s*\d|корп)\b[^)]*)\)',
    re.IGNORECASE,
)
EMOJI_RE = re.compile(r'[\U0001F300-\U0001FAFF☀-➿←-⇿]')
CTA_RE = re.compile(r'приглашаем|требуются|требуется|партнёр|партнер|подработ|ищешь|звони|пиши', re.IGNORECASE)
PRICE_RE = re.compile(r'\d[\d\s]{2,}\s*(?:₽|руб)', re.IGNORECASE)

category_patterns = {k: re.compile(v, re.IGNORECASE) for k, v in JOB_CATEGORY_KEYWORDS.items()}
emp_patterns = {k: re.compile(v, re.IGNORECASE) for k, v in EMPLOYERS.items()}


def title_of(text: str) -> str:
    return text.split("\n", 1)[0][:120]


def category_match(text: str):
    t = title_of(text)
    for cat, pat in category_patterns.items():
        if pat.search(t):
            return cat
    return None


def is_reference_mention(text: str, match_start: int, window: int = 80) -> bool:
    return bool(REFERENCE_TRIGGERS.search(text[max(0, match_start - window):match_start]))


def employer_matches(text: str) -> set:
    found = set()
    for name, pat in emp_patterns.items():
        for m in pat.finditer(text):
            if not is_reference_mention(text, m.start()):
                found.add(name)
                break
    return found


def already_visible(urls: list, ref_domains: set) -> bool:
    for u in urls:
        url = u["url"]
        if build.is_tg_resolve(url):
            continue
        if build.matches_ref_domain(url, ref_domains) or build.has_erid(url) or build.has_cpa_marker(url):
            return True
    return False


def is_platform_noise(urls: list) -> bool:
    return any(gp in u["url"] for u in urls for gp in GENERIC_PLATFORM_DOMAINS)


def is_corp_domain(employer: str, contact_type: str, contact_value: str) -> bool:
    if contact_type != "email":
        return False
    domains = CORP_DOMAINS.get(employer, ())
    val = contact_value.lower()
    return any(val.endswith("@" + d) or f"@{d}" in val for d in domains)


def is_custom_domain(contact_type: str, contact_value: str) -> bool:
    if contact_type != "email":
        return False
    dom = contact_value.lower().split("@")[-1]
    return dom not in PERSONAL_EMAIL_PROVIDERS


def has_address(titles: list) -> bool:
    return any(ADDRESS_RE.search(t) for t in titles)


def has_promo_tone(titles: list) -> bool:
    joined = " | ".join(titles)
    return bool(EMOJI_RE.search(joined) or CTA_RE.search(joined) or PRICE_RE.search(joined))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--marked", type=str, default=None,
                        help="CSV/xlsx с прошлой разметкой, колонка 'вердикт'")
    args = parser.parse_args()

    already_marked = set()
    if args.marked:
        mp = Path(args.marked)
        if mp.suffix == ".xlsx":
            wb_marked = load_workbook(mp)
            ws_marked = wb_marked.active
            header = [c.value for c in ws_marked[1]]
            vi = header.index("вердикт"); cti = header.index("contact_type"); cvi = header.index("contact_value")
            for row in ws_marked.iter_rows(min_row=2, values_only=True):
                if row[vi] and str(row[vi]).strip():
                    already_marked.add((row[cti], row[cvi]))
        else:
            for r in csv.DictReader(open(mp, encoding="utf-8")):
                if r["вердикт"].strip():
                    already_marked.add((r["contact_type"], r["contact_value"]))
        print(f"Загружено размеченных ранее: {len(already_marked)} -- они исключены из новой выгрузки")

    ref_domains = build.load_ref_domains()
    contacts_agg = {}

    for path in DUMP_PATHS:
        for line in open(path, encoding="utf-8"):
            if not line.strip():
                continue
            r = json.loads(line)
            text = r.get("text") or ""
            if not text or not r["contacts"]:
                continue

            cat = category_match(text)
            if not cat:
                continue
            emps = employer_matches(text)
            if not emps:
                continue

            vis = already_visible(r["urls"], ref_domains)
            noise = is_platform_noise(r["urls"])
            d = r["date"][:10]
            title = title_of(text)

            for c in r["contacts"]:
                key = (c["type"], c["value"])
                entry = contacts_agg.setdefault(key, {
                    "employers": set(), "categories": set(), "channels": set(),
                    "titles": set(), "message_count": 0, "already_visible": False,
                    "platform_noise": False, "first_seen": None, "last_seen": None,
                })
                entry["employers"] |= emps
                entry["categories"].add(cat)
                entry["channels"].add(r["channel"])
                entry["titles"].add(title)
                entry["message_count"] += 1
                entry["already_visible"] = entry["already_visible"] or vis
                entry["platform_noise"] = entry["platform_noise"] or noise
                if entry["first_seen"] is None or d < entry["first_seen"]:
                    entry["first_seen"] = d
                if entry["last_seen"] is None or d > entry["last_seen"]:
                    entry["last_seen"] = d

    rows = []
    for (ctype, cval), e in contacts_agg.items():
        employers_sorted = sorted(e["employers"])
        titles_list = list(e["titles"])
        is_corp = any(is_corp_domain(emp, ctype, cval) for emp in employers_sorted)
        is_custom = is_custom_domain(ctype, cval)
        addr = has_address(titles_list)
        promo = has_promo_tone(titles_list)
        junk_score = int(is_custom) + int(not promo) + int(addr)
        web_score = compute_web_score(employers_sorted, titles_list, promo)
        rows.append({
            "contact_type": ctype, "contact_value": cval,
            "employers": ", ".join(employers_sorted), "employer_count": len(employers_sorted),
            "categories": ", ".join(sorted(e["categories"])), "web_score": web_score,
            "likely_corp_hr": is_corp, "custom_domain": is_custom,
            "has_address": addr, "has_promo_tone": promo, "junk_score": junk_score,
            "channel_count": len(e["channels"]), "message_count": e["message_count"],
            "already_visible_to_pipeline": e["already_visible"], "platform_noise": e["platform_noise"],
            "likely_aggregator": len(employers_sorted) >= EMPLOYER_COUNT_NOISE_THRESHOLD,
            "first_seen": e["first_seen"], "last_seen": e["last_seen"],
            "sample_titles": " | ".join(titles_list[:3]),
            "channels": ", ".join(sorted(e["channels"])[:10]),
        })

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    review_rows = [r for r in rows if not r["platform_noise"] and not r["already_visible_to_pipeline"]
                   and not r["likely_aggregator"] and not r["has_address"] and not r["custom_domain"]
                   and (r["contact_type"], r["contact_value"]) not in already_marked]
    review_rows.sort(key=lambda r: (-r["web_score"], -r["channel_count"]))

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["вердикт", "employers", "employer_count", "categories", "web_score",
                  "contact_type", "contact_value", "junk_score", "has_promo_tone",
                  "channel_count", "message_count", "first_seen", "last_seen",
                  "sample_titles", "channels"]
    wb = Workbook()
    ws = wb.active
    ws.title = "review"
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Arial")
    for r in review_rows:
        ws.append([r.get(k, "") for k in fieldnames])

    dv = DataValidation(type="list", formula1='"веб,не_веб,не_уверен"', allow_blank=True)
    dv.errorTitle = "Недопустимое значение"
    dv.error = "Выбери значение из списка"
    ws.add_data_validation(dv)
    dv.add(f"A2:A{max(len(review_rows) + 1, 2)}")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["N"].width = 60  # sample_titles
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_XLSX)
    print(f"XLSX с дропдауном -> {OUTPUT_XLSX.relative_to(BASE_DIR)}")

    print(f"Уникальных контактов всего: {len(rows)}")
    print(f"Площадки-шумогенераторы: {sum(1 for r in rows if r['platform_noise'])}")
    print(f"Вероятные агрегаторы: {sum(1 for r in rows if r['likely_aggregator'] and not r['platform_noise'])}")
    print(f"Уже видны пайплайну: {sum(1 for r in rows if r['already_visible_to_pipeline'])}")
    print(f"На ручную фильтрацию: {len(review_rows)}")
    for score in range(4):
        print(f"  junk_score={score}: {sum(1 for r in review_rows if r['junk_score']==score)}")
    print(f"\nJSON -> {OUTPUT_JSON.relative_to(BASE_DIR)}")


if __name__ == "__main__":
    main()

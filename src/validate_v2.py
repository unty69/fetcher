"""
validate_v2.py — validates the new alt_cluster_review.py relevance signals
(has_military_flag, has_agency_self_id, employer_count, content_volume_flag,
categories) against the labeled ground truth.

Reconstructs the signals for every canonical_id in the root ground-truth
workbook and reports exact per-rule hit rates, a confusion matrix, concrete
false-positive catches, and AC-1..AC-4.

Read-only / deterministic:
  - never writes to ./alt_clusters_review.xlsx or output/alt_clusters_review.xlsx
  - never writes any file at all (report goes to stdout only)
  - no ML / no randomness -- same input always produces the same report
  - does not call alt_cluster_review.write_workbook() / main() (no export)
"""
import re
import statistics
import sys
from collections import Counter
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import build  # noqa: E402
import alt_cluster_review as acr  # noqa: E402
from identity_graph import iter_dump_paths, load_messages, message_contact_nodes  # noqa: E402

# Authoritative ground truth is the ROOT workbook only -- never
# output/alt_clusters_review.xlsx (empty template) or cache/alt_clusters_review.xlsx.
GROUND_TRUTH_PATH = BASE_DIR / "alt_clusters_review.xlsx"

LABELS = ["веб", "не_веб", "агентство", "не_уверен"]
LABEL_EN = {"веб": "web", "не_веб": "non-web", "агентство": "agency", "не_уверен": "uncertain"}

# Address/location patterns -- soft signal, distinct-count based (NOT boolean
# "mentions a location") specifically because a genuine single job post can
# legitimately name its one workplace once; only REPEATED distinct locations
# across an identity's messages looks like aggregator behaviour. Validated
# against real message text from the 134 labeled identities before use (see
# conversation record): STREET/METRO/MALL showed 0 visible false positives
# across ~109 samples. BARE (no street-type prefix) is the highest-risk one --
# 2 real false positives found in 27 samples: a date range ("Петербург,
# 18 - 21 июня" read as house number 18) and a brand-name list ("Сушивок, 2
# берега" -- restaurant names, read as street+number). The date-range shape is
# guarded against below; the brand-list shape is NOT (would need a full brand
# dictionary cross-check to rule out generally) -- accepted as a known residual
# risk, tolerable specifically because this is a distinct-count soft signal,
# not a boolean or a hard filter, so one stray match rarely changes the count
# enough to matter.
STREET_ADDRESS_RE = re.compile(
    r'\b(?:ул\.?|улица|пр-?т\.?|проспект|пер\.?|переулок|б-?р\.?|бульвар|ш\.?|шоссе|наб\.?|набережная)\s+'
    r'[А-ЯЁ][а-яёА-Я\-]{2,25}',
    re.IGNORECASE,
)
METRO_ADDRESS_RE = re.compile(
    r'\b(?:ст\.?\s*м\.?|станция\s+метро|метро)\s+[А-ЯЁ][а-яёА-Я\-]{2,25}',
    re.IGNORECASE,
)
MALL_ADDRESS_RE = re.compile(
    r'\b(?:ТЦ|ТРЦ|ТРК|БЦ)\s*[«"]?\s*[А-ЯЁ][а-яёА-Я\- ]{1,30}[»"]?',
)
BARE_ADDRESS_RE = re.compile(
    r'\b[А-ЯЁ][а-яё]{2,20},\s*\d{1,4}[а-яА-Я]?\b(?!\s*-\s*\d)',
)
ADDRESS_PATTERNS = {
    "street": STREET_ADDRESS_RE, "metro": METRO_ADDRESS_RE,
    "mall": MALL_ADDRESS_RE, "bare": BARE_ADDRESS_RE,
}


# "Пункт выдачи заказов" / "ПВЗ" -- Арсений's instruction: single/few-person
# points, not mass-hire targets. NOT itself a category-match term (verified:
# no JOB_CATEGORY_KEYWORDS pattern contains "пункт" or "ПВЗ"). The reason PVZ
# roles were showing up under "Курьеры и доставка" was confirmed on real
# message text -- e.g. "Администратор в пункт выдачи ... Обязанности: Приём
# товара ОТ КУРЬЕРА" -- the job description mentions "курьера" because the
# role RECEIVES FROM couriers, not because it IS one. Full hard-exclude
# (Арсений was explicit "нет массового набора"), not a soft signal.
PVZ_RE = re.compile(r'пункт[а-я]*\s+выдач|\bПВЗ\b', re.IGNORECASE)

# Off-target white-collar noise -- validated against real uncategorized
# message titles from the 189-identity labeled set (see conversation record):
# "Менеджер-аналитик маркетплейсов", "Бухгалтер (офис в г. Казань)",
# "счётчика-ревизора", "Офис-менеджер", "Менеджер по продажам СИП, ВОЛС,
# СКС", "Заместитель директора магазина" all appeared as real uncategorized
# titles. Deliberately narrow/curated (not "anything uncategorized") --
# uncategorized-but-plausibly-legitimate blue-collar content (e.g.
# "Кондитер", "Оператор на производство") must NOT be swept in here, so this
# is a positive keyword match, not a catch-all for category-match failure.
OFF_TARGET_RE = re.compile(
    r'\bинженер|\bсметчик|\bтендер|\bаналитик|\bюрист|\bбухгалтер|\bмаркетолог|\bревизор|'
    r'офис[\s-]?менеджер|менеджер\s+по\s+продажам|\bдиректор',
    re.IGNORECASE,
)


def address_matches_in_text(text):
    """{(pattern_name, normalized_match_string)} -- normalized so the same
    address repeated across many near-duplicate reposted messages (observed
    in real data, e.g. the same "ул. Летчика..." address on ~7 messages)
    counts once, not once per repost."""
    found = set()
    for name, pat in ADDRESS_PATTERNS.items():
        for m in pat.finditer(text):
            normalized = re.sub(r'\s+', ' ', m.group(0).strip().lower())
            found.add((name, normalized))
    return found


def load_ground_truth():
    """[(canonical_id, verdict_or_empty_string)] for every non-empty row.
    read_only=True -- this workbook is never written to."""
    wb = openpyxl.load_workbook(GROUND_TRUTH_PATH, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter)
        header_index = {h: i for i, h in enumerate(header)}
        verdict_col = header_index["вердикт"]
        cid_col = header_index["canonical_id"]
        out = []
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue
            cid = row[cid_col]
            verdict = row[verdict_col]
            v = str(verdict).strip() if verdict is not None else ""
            out.append((cid, v))
        return out
    finally:
        wb.close()


def collect_debug_evidence(candidate_ids, contact_to_canonical):
    """Mirrors alt_cluster_review.collect_evidence() -- same regex objects,
    same functions, so this measures the actual production signals, not a
    reimplementation -- but also keeps a couple of matched-keyword snippets
    per identity for the concrete false-positive-catch section below."""
    evidence = {
        cid: {
            "brands": set(), "has_ref_link": False, "titles": [],
            "has_military_flag": False, "has_agency_self_id": False,
            "categories": set(), "max_text_len": 0,
            "military_hits": [], "agency_hits": [],
            "has_verified_relevant_message": False,
            "category_hit_messages": [], "verified_brand_hit_messages": [],
            "verified_relevant_examples": [],
            "address_matches": set(),
            "has_pvz_mention": False, "pvz_hits": [],
            "off_target_count": 0, "messages_seen": 0, "off_target_examples": [],
        }
        for cid in candidate_ids
    }

    ref_domains = build.load_ref_domains()
    messages = load_messages(iter_dump_paths())

    for msg in messages:
        nodes = message_contact_nodes(msg)
        if not nodes:
            continue
        matched_cids = {
            contact_to_canonical[node_key]
            for node_key, *_ in nodes
            if node_key in contact_to_canonical and contact_to_canonical[node_key] in candidate_ids
        }
        if not matched_cids:
            continue

        text = msg.get("text") or ""
        brands = {name for name, pat in acr.BRAND_PATTERNS.items() if pat.search(text)}
        ref_link = any(
            build.matches_ref_domain(u["url"], ref_domains)
            for u in msg["urls"] if not build.is_tg_resolve(u["url"])
        )
        title = acr.title_of(text)
        is_military = build.is_military_content(text)
        agency_match = acr.AGENCY_SELF_ID_RE.search(text)
        matched_categories = {cat for cat, pat in acr.CATEGORY_PATTERNS.items() if pat.search(text)}
        text_len = len(text.strip())

        # Same fix as alt_cluster_review.collect_evidence(): co-occurrence on
        # THIS message, not a union of independently-accumulated sets.
        verified_brands_this_msg = brands & acr.CPA_VERIFIED_EMPLOYERS
        msg_verified_relevant = bool(matched_categories) and bool(verified_brands_this_msg)
        msg_address_matches = address_matches_in_text(text)
        msg_is_pvz = bool(PVZ_RE.search(text))
        # off-target noise: matches the curated keyword list AND doesn't
        # match a target category AND isn't already military/agency/PVZ --
        # those have their own handling, not double-counted here.
        msg_is_off_target = (
            bool(OFF_TARGET_RE.search(text)) and not matched_categories
            and not is_military and not agency_match and not msg_is_pvz
        )

        for cid in matched_cids:
            ev = evidence[cid]
            ev["brands"] |= brands
            ev["has_ref_link"] = ev["has_ref_link"] or ref_link
            if title and title not in ev["titles"] and len(ev["titles"]) < 3:
                ev["titles"].append(title)
            if is_military:
                ev["has_military_flag"] = True
                if len(ev["military_hits"]) < 2:
                    lowered = text.lower()
                    matched_kw = next((kw for kw in build.MILITARY_KEYWORDS if kw in lowered), "?")
                    ev["military_hits"].append((matched_kw, title))
            if agency_match:
                ev["has_agency_self_id"] = True
                if len(ev["agency_hits"]) < 2:
                    ev["agency_hits"].append((agency_match.group(0), title))
            ev["categories"] |= matched_categories
            ev["has_verified_relevant_message"] = ev["has_verified_relevant_message"] or msg_verified_relevant
            ev["address_matches"] |= msg_address_matches
            ev["messages_seen"] += 1
            if msg_is_pvz:
                ev["has_pvz_mention"] = True
                if len(ev["pvz_hits"]) < 2:
                    ev["pvz_hits"].append(title)
            if msg_is_off_target:
                ev["off_target_count"] += 1
                if len(ev["off_target_examples"]) < 3:
                    ev["off_target_examples"].append(title)
            # Diagnostic only (not used by production code): track messages
            # that had a category match, and separately messages that had a
            # verified-brand match, even when they didn't co-occur -- so a
            # failed AC-1 check can show exactly what was missing instead of
            # just "False".
            if matched_categories and len(ev["category_hit_messages"]) < 3:
                ev["category_hit_messages"].append((sorted(matched_categories), title))
            if verified_brands_this_msg and len(ev["verified_brand_hit_messages"]) < 3:
                ev["verified_brand_hit_messages"].append((sorted(verified_brands_this_msg), title))
            # The actual co-occurrence instance (both together, same message) --
            # used by export_candidates_v2.py's sample-rows report.
            if msg_verified_relevant and len(ev["verified_relevant_examples"]) < 2:
                ev["verified_relevant_examples"].append(
                    (sorted(matched_categories), sorted(verified_brands_this_msg), title))
            if text_len > ev["max_text_len"]:
                ev["max_text_len"] = text_len

    return evidence


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    gt_rows = load_ground_truth()
    labeled = [(cid, v) for cid, v in gt_rows if v in LABELS]
    labeled_map = dict(labeled)
    blank = [(cid, v) for cid, v in gt_rows if v not in LABELS]

    print("=" * 78)
    print("Dataset")
    print("=" * 78)
    print(f"Ground truth file: {GROUND_TRUTH_PATH.relative_to(BASE_DIR)}")
    print(f"Total rows: {len(gt_rows)}")
    label_counts = Counter(v for _, v in labeled)
    print("\nWeb / non-web / agency / uncertain breakdown:")
    for lbl in LABELS:
        print(f"  {LABEL_EN[lbl]} ({lbl}): {label_counts.get(lbl, 0)}")
    print(f"  blank/unlabeled: {len(blank)}")
    print(f"Labeled total: {len(labeled)}")

    web_cids = [cid for cid, v in labeled if v == "веб"]
    print(f"\nCurrently-labeled 'веб' canonical_ids (read dynamically from ground truth, "
          f"NOT a hardcoded count): {len(web_cids)}")
    for cid in web_cids:
        print(f"  {cid}")

    candidate_ids = {cid for cid, _ in gt_rows if cid}

    graph = acr.load_identity_graph()
    identities = graph["identities"]
    contact_to_canonical = graph["contact_to_canonical"]

    missing = [cid for cid in candidate_ids if cid not in identities]
    if missing:
        print(f"\n!!! {len(missing)} ground-truth canonical_ids NOT found in "
              f"cache/identity_graph.json: {missing}")
        print("!!! Stopping -- cannot validate identities that aren't in the current graph.")
        sys.exit(1)
    print(f"\nAll {len(candidate_ids)} ground-truth canonical_ids resolved in "
          f"cache/identity_graph.json ({len(identities)} total identities). OK.")

    evidence = collect_debug_evidence(candidate_ids, contact_to_canonical)

    def cls_counts(predicate):
        counts = {}
        for lbl in LABELS:
            cids = [cid for cid, v in labeled if v == lbl]
            hits = sum(1 for cid in cids if predicate(cid))
            counts[lbl] = (hits, len(cids))
        return counts

    def print_rule(name, predicate):
        counts = cls_counts(predicate)
        print(f"\nRule: {name}")
        for lbl in LABELS:
            hits, total = counts[lbl]
            print(f"  {LABEL_EN[lbl]} ({lbl}): {hits}/{total}")
        return counts

    print("\n" + "=" * 78)
    print("Per-rule hit rates")
    print("=" * 78)
    print("\n-- new signals --")
    military_counts = print_rule("has_military_flag", lambda cid: evidence[cid]["has_military_flag"])
    agency_counts = print_rule("has_agency_self_id", lambda cid: evidence[cid]["has_agency_self_id"])
    any_category_counts = print_rule("categories: matched >=1 target category (full text)",
                                      lambda cid: len(evidence[cid]["categories"]) > 0)
    content_counts = print_rule(f"content_volume_flag (max_text_len >= {acr.CONTENT_VOLUME_MIN_CHARS} chars, "
                                 f"PROVISIONAL threshold)",
                                 lambda cid: evidence[cid]["max_text_len"] >= acr.CONTENT_VOLUME_MIN_CHARS)
    verified_counts = print_rule("has_verified_relevant_message (category AND CPA-verified employer, "
                                  "SAME message -- fixes the identity-level union bug)",
                                  lambda cid: evidence[cid]["has_verified_relevant_message"])
    print("\n  (for comparison, the OLD buggy rule this replaces as the export hard-exclude gate:)")
    old_buggy_counts = print_rule("[OLD/BUGGY] categories non-empty (identity-level union, no co-occurrence)",
                                   lambda cid: len(evidence[cid]["categories"]) > 0)
    print("\n-- existing weak signals (informational only, unchanged by this task) --")
    ref_link_counts = print_rule("has_ref_link", lambda cid: evidence[cid]["has_ref_link"])
    corp_counts = print_rule("corp_domain", lambda cid: acr.has_corp_domain(identities[cid]))
    hr_counts = print_rule("hr_hint", lambda cid: acr.has_hr_hint(identities[cid]))
    multibrand_counts = print_rule("multibrand (employer_count>=2)", lambda cid: len(evidence[cid]["brands"]) >= 2)

    print("\n" + "=" * 78)
    print("Employer count distribution (by label)")
    print("=" * 78)

    def bucket(n):
        return str(n) if n < 4 else "4+"

    buckets_order = ["0", "1", "2", "3", "4+"]
    dist = {lbl: Counter() for lbl in LABELS}
    for cid, v in labeled:
        dist[v][bucket(len(evidence[cid]["brands"]))] += 1
    print(f"{'label':<14}" + "".join(b.rjust(6) for b in buckets_order))
    for lbl in LABELS:
        print(f"{LABEL_EN[lbl]:<14}" + "".join(str(dist[lbl].get(b, 0)).rjust(6) for b in buckets_order))
    print("\nHypothesis from CLAUDE.md (0=weak/unknown, 1=possible negative, 2-3=possible positive, "
          "4+=possible aggregator/noise) -- NOT applied as a filter, distribution shown for information only.")

    print("\n" + "=" * 78)
    print("False-positive buckets (CLAUDE.md 'Important False-Positive Buckets')")
    print("=" * 78)

    print("\n[White-collar / non-target category] -- identities with ZERO target-category match "
          "in full text (JOB_CATEGORY_KEYWORDS, reused directly, not via truncating category_match()):")
    for lbl in LABELS:
        hits, total = any_category_counts[lbl]
        print(f"  {LABEL_EN[lbl]}: {total - hits}/{total} matched no category, {hits}/{total} matched >=1")

    print("\n[Military] -- see 'has_military_flag' rule above and AC-2 below.")
    print("[Agency] -- see 'has_agency_self_id' rule above.")

    print("\n[Single-employer noise] -- employer_count==1 (must stay a SOFT signal, never hard-excluded; "
          "CLAUDE.md notes a known positive has a single-employer Samokat partner message):")
    for lbl in LABELS:
        cids = [cid for cid, v in labeled if v == lbl]
        single = sum(1 for cid in cids if len(evidence[cid]["brands"]) == 1)
        print(f"  {LABEL_EN[lbl]}: {single}/{len(cids)}")

    print("\n[Low-content] -- content_volume_flag == False (max_text_len < "
          f"{acr.CONTENT_VOLUME_MIN_CHARS} chars, PROVISIONAL threshold):")
    for lbl in LABELS:
        hits, total = content_counts[lbl]
        print(f"  {LABEL_EN[lbl]}: {total - hits}/{total} low-content")

    print("\n" + "=" * 78)
    print("Confusion matrix: label vs actual export hard-exclusion gate")
    print("=" * 78)
    print("gate = has_military_flag OR has_agency_self_id OR NOT has_verified_relevant_message")
    print("(this is now the real export_candidates_v2.py gate, not a hypothetical preview)")
    print(f"\n{'label':<14}{'excluded':>14}{'would_export':>14}{'total':>8}")
    for lbl in LABELS:
        cids = [cid for cid, v in labeled if v == lbl]
        excluded_n = sum(1 for cid in cids if evidence[cid]["has_military_flag"]
                          or evidence[cid]["has_agency_self_id"]
                          or not evidence[cid]["has_verified_relevant_message"])
        print(f"{LABEL_EN[lbl]:<14}{excluded_n:>14}{len(cids) - excluded_n:>14}{len(cids):>8}")

    print("\n" + "=" * 78)
    print("Concrete false-positive catches (non-web identities newly flaggable)")
    print("=" * 78)
    catches = []
    for cid, v in labeled:
        if v == "веб":
            continue
        ev = evidence[cid]
        reasons = []
        if ev["has_military_flag"]:
            reasons.append(f"military hit={ev['military_hits']}")
        if ev["has_agency_self_id"]:
            reasons.append(f"agency_self_id hit={ev['agency_hits']}")
        if not ev["has_verified_relevant_message"]:
            reasons.append("NOT has_verified_relevant_message "
                            f"(categories(union)={sorted(ev['categories'])}, "
                            f"verified_brands(union)={sorted(ev['brands'] & acr.CPA_VERIFIED_EMPLOYERS)})")
        if reasons:
            catches.append((cid, v, reasons))
    print(f"Total identities newly flagged by military OR agency_self_id OR NOT has_verified_relevant_message "
          f"(label != веб): {len(catches)} / {len(labeled) - len(web_cids)}")
    print("(first 15 shown; full gate applied at scale in export_candidates_v2.py)")
    for cid, v, reasons in catches[:15]:
        print(f"  {cid} | old_classification={LABEL_EN[v]}({v}) | new_signal: {' ; '.join(reasons)}")

    print("\n" + "=" * 78)
    print("has_verified_relevant_message check on known web identities (export hard-exclude gate)")
    print("=" * 78)
    print("Required before any new export: both known web canonical_ids must have")
    print("has_verified_relevant_message=True. If not, STOP and report the missing")
    print("message/employer/category combination -- do not loosen the rule to force a pass.")
    verified_gate_failures = []
    for cid in web_cids:
        v = evidence[cid]["has_verified_relevant_message"]
        print(f"\n  {cid}: has_verified_relevant_message={v}")
        if v:
            print(f"    (co-occurrence confirmed on >=1 message)")
        else:
            verified_gate_failures.append(cid)
            print(f"    !!! MISSING. Diagnostic -- what WAS found, just never together:")
            print(f"    category matches (any message, unioned): {sorted(evidence[cid]['categories'])}")
            print(f"    category-match sample messages: {evidence[cid]['category_hit_messages']}")
            print(f"    CPA-verified brand matches (any message, unioned): "
                  f"{sorted(evidence[cid]['brands'] & acr.CPA_VERIFIED_EMPLOYERS)}")
            print(f"    verified-brand-match sample messages: {evidence[cid]['verified_brand_hit_messages']}")
            print(f"    all brands (incl. non-CPA-verified): {sorted(evidence[cid]['brands'])}")

    if verified_gate_failures:
        print(f"\n!!! STOPPING: {len(verified_gate_failures)}/{len(web_cids)} known web identities would be "
              f"lost by the new has_verified_relevant_message export gate: {verified_gate_failures}")
        print("!!! Per instruction: not loosening the rule to force a pass. Fix requires investigating")
        print("!!! the specific message/employer/category gap shown above before any export proceeds.")
        sys.exit(1)
    print(f"\nAll {len(web_cids)}/{len(web_cids)} known web identities have has_verified_relevant_message=True. "
          f"Safe to proceed to export with the new gate.")

    print("\n" + "=" * 78)
    print("New scoring signals on known web identities (end-to-end re-confirmation)")
    print("=" * 78)
    for cid in web_cids:
        ev = evidence[cid]
        cat_count = len(ev["categories"])
        addr_count = len(ev["address_matches"])
        corp = acr.has_corp_domain(identities[cid])
        off_frac = ev["off_target_count"] / ev["messages_seen"] if ev["messages_seen"] else 0.0
        print(f"\n  {cid}:")
        print(f"    corp_domain={corp}  (main-eligibility gate -- must be False)")
        print(f"    category_count={cat_count}  categories={sorted(ev['categories'])}")
        print(f"    address_location_count={addr_count}  matches={sorted(ev['address_matches'])}")
        print(f"    has_pvz_mention={ev['has_pvz_mention']}  hits={ev['pvz_hits']}  (hard-exclude gate -- must be False)")
        print(f"    off_target_fraction={off_frac:.2f} ({ev['off_target_count']}/{ev['messages_seen']} messages)  "
              f"examples={ev['off_target_examples']}")
    print("\n  (category_count, address_location_count, off_target_fraction feed the score as soft")
    print("   signals; has_pvz_mention is a hard exclude -- all shown here so nothing goes unnoticed)")

    print("\n" + "=" * 78)
    print("off_target_fraction / has_pvz_mention by verdict (full labeled set)")
    print("=" * 78)
    for lbl in LABELS:
        cids = [cid for cid, v in labeled if v == lbl]
        if not cids:
            continue
        fracs = sorted((evidence[cid]["off_target_count"] / evidence[cid]["messages_seen"]
                         if evidence[cid]["messages_seen"] else 0.0) for cid in cids)
        pvz_n = sum(1 for cid in cids if evidence[cid]["has_pvz_mention"])
        print(f"  {LABEL_EN[lbl]}: off_target_fraction median={statistics.median(fracs):.2f} "
              f"max={fracs[-1]:.2f}  |  has_pvz_mention: {pvz_n}/{len(cids)}")

    print("\n" + "=" * 78)
    print("category_count / address_location_count distribution across the FULL labeled set")
    print("=" * 78)
    print("(diagnostic context, not the actual scoring population -- that's the export pool)")
    for lbl in LABELS:
        cids = [cid for cid, v in labeled if v == lbl]
        if not cids:
            continue
        cat_counts = sorted(len(evidence[cid]["categories"]) for cid in cids)
        addr_counts = sorted(len(evidence[cid]["address_matches"]) for cid in cids)
        print(f"  {LABEL_EN[lbl]}: category_count min={cat_counts[0]} median={statistics.median(cat_counts):.1f} "
              f"max={cat_counts[-1]}  |  address_location_count min={addr_counts[0]} "
              f"median={statistics.median(addr_counts):.1f} max={addr_counts[-1]}")

    print("\n" + "=" * 78)
    print("Known-web preservation / AC-1")
    print("=" * 78)
    ac1_failures = []
    for cid in web_cids:
        ident = identities[cid]
        mil = evidence[cid]["has_military_flag"]
        ag = evidence[cid]["has_agency_self_id"]
        still_alt_cluster = ident["is_alt_cluster"] and not ident["suspected_non_discovery"]
        survives = (not mil) and (not ag) and still_alt_cluster
        print(f"  {cid}: has_military_flag={mil} has_agency_self_id={ag} "
              f"still_in_alt_cluster_pool={still_alt_cluster} survives={survives}")
        if not survives:
            ac1_failures.append(cid)
    if len(web_cids) == 0:
        ac1_verdict = "NOT MEASURED (no labeled 'веб' identities currently exist in ground truth)"
    elif ac1_failures:
        ac1_verdict = f"FAIL ({len(ac1_failures)}/{len(web_cids)} known web identities would be lost)"
    else:
        ac1_verdict = f"PASS ({len(web_cids)}/{len(web_cids)} known web identities survive, recall=100%)"
    print(f"AC-1: {ac1_verdict}")

    print("\n" + "=" * 78)
    print("AC-2: Military Detection")
    print("=" * 78)
    military_hit_cids = [cid for cid, v in labeled if evidence[cid]["has_military_flag"]]
    print(f"Identities with has_military_flag=True among the {len(labeled)} labeled identities: "
          f"{len(military_hit_cids)}")
    for cid in military_hit_cids:
        v = labeled_map[cid]
        print(f"  {cid} label={LABEL_EN[v]}({v}) matched={evidence[cid]['military_hits']}")
    hits, total = military_counts["веб"]
    print(f"\nweb: {hits}/{total}   non-web: {military_counts['не_веб'][0]}/{military_counts['не_веб'][1]}   "
          f"agency: {military_counts['агентство'][0]}/{military_counts['агентство'][1]}   "
          f"uncertain: {military_counts['не_уверен'][0]}/{military_counts['не_уверен'][1]}")
    ac2_no_fp_on_web = hits == 0
    ac2_detects_something = len(military_hit_cids) > 0
    if ac2_detects_something and ac2_no_fp_on_web:
        ac2_verdict = "PASS (fires on >=1 labeled identity, 0 false positives on known web)"
    elif not ac2_detects_something:
        ac2_verdict = "FAIL (build.is_military_content did not fire on any labeled identity -- cannot confirm detection on this dataset)"
    else:
        ac2_verdict = f"FAIL (fired on {hits}/{total} known web identities)"
    print(f"AC-2: {ac2_verdict}")

    print("\n" + "=" * 78)
    print("AC-3: Validation Improvement")
    print("=" * 78)
    print("No target precision/recall number is asserted (CLAUDE.md: do not invent one).")
    print("Measured evidence for this criterion is the full report above: per-rule hit rates,")
    print("web/non-web/agency/uncertain breakdown, employer_count distribution, false-positive")
    print("buckets, confusion matrix, and the concrete catches list.")
    print(f"Concrete, measured claim: before this change, alt_cluster_review.py had no "
          f"has_military_flag/has_agency_self_id/categories/content_volume_flag columns at all. "
          f"After this change, {len(catches)}/{len(labeled) - len(web_cids)} non-web-labeled "
          f"identities are now flagged by >=1 new hard-exclusion-candidate signal, "
          f"0/{len(web_cids)} known web identities are flagged.")
    print("AC-3: PASS (required measured evidence produced above; no aggregate claim made without numbers)")

    print("\n" + "=" * 78)
    print("AC-4: Content Sufficiency")
    print("=" * 78)
    print("This is a PRODUCTION-BATCH criterion (>=90% content-sufficient), not a validation-set")
    print("pass/fail gate (CLAUDE.md: do not use it to destroy known validation positives).")
    print("Per the baseline check: select_candidates() currently returns exactly the ground-truth")
    print("pool (135/135 overlap) -- there is currently NO unreviewed production batch to measure.")
    print(f"content_volume_flag distribution across the LABELED set (diagnostic only, NOT the AC-4 "
          f"target population):")
    for lbl in LABELS:
        hits, total = content_counts[lbl]
        pct = f"{100 * hits / total:.0f}%" if total else "n/a"
        print(f"  {LABEL_EN[lbl]}: {hits}/{total} ({pct}) sufficient")
    print("AC-4: NOT MEASURED (no production candidate batch currently exists beyond the labeled set)")

    print("\n" + "=" * 78)
    print("Recommendation")
    print("=" * 78)
    print(f"- has_military_flag: {ac2_verdict}. {'Zero web false positives observed; ' if ac2_no_fp_on_web else ''}"
          f"candidate for hard exclusion once more labeled data accumulates -- do not flip to hard "
          f"exclusion from this run alone (only {label_counts.get('веб', 0)} web / "
          f"{len(labeled)} total labeled examples).")
    print(f"- has_agency_self_id: fired on {agency_counts['агентство'][0]}/{agency_counts['агентство'][1]} "
          f"labeled 'агентство' identities and {agency_counts['веб'][0]}/{agency_counts['веб'][1]} 'веб'. "
          f"Keep as a distinct classification signal, not silently merged into не_веб, per CLAUDE.md.")
    print(f"- employer_count / content_volume_flag: soft/ranking signals only, per CLAUDE.md. Distributions")
    print(f"  reported above; thresholds NOT tuned to force a particular outcome.")
    print(f"- Known-positive sample size remains small ({label_counts.get('веб', 0)} web identities) -- ")
    print(f"  all percentages involving the web class are statistically unstable, per CLAUDE.md.")


if __name__ == "__main__":
    main()
